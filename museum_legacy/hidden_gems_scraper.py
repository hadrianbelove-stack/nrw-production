#!/usr/bin/env python3
"""
Unified Hidden Gems Scraper
---------------------------
Finds indie films with digital release dates (Type 4) but not on major streaming/rental platforms.
These are self-distributed films on Vimeo, YouTube, Patreon, personal sites, etc.

🎬 UNIFIED SCRAPER - All platforms in one tool!

Features:
- TMDB discovery (Type 4 release, no providers)
- Live scraping from Vimeo On Demand, YouTube Movies, Patreon creators, Letterboxd
- Curation-optimized exports
- Cross-platform deduplication

Usage Examples:
    # Standard TMDB discovery
    python3 hidden_gems_scraper.py --days 30
    python3 hidden_gems_scraper.py --month 2025-12

    # Live scraping (uses standalone scrapers)
    python3 hidden_gems_scraper.py --scrape-all --min-runtime 60
    python3 hidden_gems_scraper.py --scrape-vimeo --scrape-patreon --max-pages 5
    python3 hidden_gems_scraper.py --scrape-youtube --max-pages 5 --no-headless

    # Hybrid: TMDB + Live scraping
    python3 hidden_gems_scraper.py --days 7 --scrape-vimeo --export-for-curation

    # Curation workflow
    python3 hidden_gems_scraper.py --scrape-all --export-for-curation --output my_gems
"""

import argparse
import csv
import json
import logging
import os
import time
import re
from datetime import datetime, timedelta
from typing import Optional, Dict, List, Tuple
from urllib.parse import urlparse
import requests

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Configure logging
logger = logging.getLogger(__name__)

# TMDB API
TMDB_API_KEY = os.environ.get('TMDB_API_KEY', '')
if not TMDB_API_KEY:
    try:
        import yaml
        with open('config.yaml', 'r') as f:
            config = yaml.safe_load(f)
            TMDB_API_KEY = config.get('api', {}).get('tmdb_api_key', '')
    except:
        pass
TMDB_BASE = 'https://api.themoviedb.org/3'

# Import standalone scrapers (refactored to use shared base class)
try:
    from patreon_film_scraper import PatreonFilmScraper
    PATREON_AVAILABLE = True
except ImportError:
    logger.warning("PatreonFilmScraper not available")
    PATREON_AVAILABLE = False

try:
    from vimeo_ondemand_scraper import VimeoOnDemandScraper
    VIMEO_AVAILABLE = True
except ImportError:
    logger.warning("VimeoOnDemandScraper not available")
    VIMEO_AVAILABLE = False

try:
    from youtube_paid_scraper import YouTubePaidScraper
    YOUTUBE_AVAILABLE = True
except ImportError:
    logger.warning("YouTubePaidScraper not available")
    YOUTUBE_AVAILABLE = False

try:
    from letterboxd_scraper import LetterboxdScraper
    LETTERBOXD_AVAILABLE = True
except ImportError:
    logger.warning("LetterboxdScraper not available")
    LETTERBOXD_AVAILABLE = False


# ===== DATE UTILITY FUNCTIONS =====

def parse_relative_date(text: str) -> Optional[str]:
    """Parse relative date strings like '2 days ago', '1 week ago' into YYYY-MM-DD format."""
    if not text:
        return None

    text = text.lower().strip()
    today = datetime.now()

    # Try to parse as ISO date first
    iso_patterns = [
        r'(\d{4})-(\d{2})-(\d{2})',  # 2025-01-15
        r'(\d{2})/(\d{2})/(\d{4})',  # 01/15/2025
    ]
    for pattern in iso_patterns:
        match = re.search(pattern, text)
        if match:
            try:
                if '/' in text:
                    return f"{match.group(3)}-{match.group(1)}-{match.group(2)}"
                return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
            except:
                pass

    # Parse relative dates
    relative_patterns = [
        (r'(\d+)\s*(?:second|sec)s?\s*ago', lambda m: timedelta(seconds=int(m.group(1)))),
        (r'(\d+)\s*(?:minute|min)s?\s*ago', lambda m: timedelta(minutes=int(m.group(1)))),
        (r'(\d+)\s*(?:hour|hr)s?\s*ago', lambda m: timedelta(hours=int(m.group(1)))),
        (r'(\d+)\s*(?:day)s?\s*ago', lambda m: timedelta(days=int(m.group(1)))),
        (r'(\d+)\s*(?:week)s?\s*ago', lambda m: timedelta(weeks=int(m.group(1)))),
        (r'(\d+)\s*(?:month)s?\s*ago', lambda m: timedelta(days=int(m.group(1)) * 30)),
        (r'(\d+)\s*(?:year)s?\s*ago', lambda m: timedelta(days=int(m.group(1)) * 365)),
        (r'yesterday', lambda m: timedelta(days=1)),
        (r'today', lambda m: timedelta(days=0)),
    ]

    for pattern, delta_fn in relative_patterns:
        match = re.search(pattern, text)
        if match:
            try:
                delta = delta_fn(match)
                result_date = today - delta
                return result_date.strftime('%Y-%m-%d')
            except:
                pass

    return None


def is_within_date_range(date_str: str, start_date: str, end_date: str) -> bool:
    """Check if a date string falls within the specified range."""
    if not date_str:
        return False
    try:
        date = datetime.strptime(date_str, '%Y-%m-%d')
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        return start <= date <= end
    except:
        return False


def get_dated_output_prefix() -> str:
    """Generate dated, numbered output path in hidden_gems_output/ folder.

    Returns path like: hidden_gems_output/2026-02-09_001_hidden_gems
    """
    output_dir = 'hidden_gems_output'
    os.makedirs(output_dir, exist_ok=True)

    today = datetime.now().strftime('%Y-%m-%d')

    # Find existing files for today to determine next number
    existing_nums = []
    for filename in os.listdir(output_dir):
        if filename.startswith(today) and '_hidden_gems' in filename:
            # Extract number from pattern like 2026-02-09_001_hidden_gems.xlsx
            parts = filename.split('_')
            if len(parts) >= 2:
                try:
                    num = int(parts[1])
                    existing_nums.append(num)
                except ValueError:
                    pass

    next_num = max(existing_nums, default=0) + 1

    return os.path.join(output_dir, f"{today}_{next_num:03d}_hidden_gems")


def calculate_date_range(month: Optional[str] = None, days: Optional[int] = None) -> Tuple[str, str]:
    """Calculate start and end dates for filtering."""
    if month:
        # e.g., "2025-12" -> 2025-12-01 to 2025-12-31
        year, mon = month.split('-')
        start_date = f"{month}-01"
        if int(mon) == 12:
            end_date = f"{int(year)+1}-01-01"
        else:
            end_date = f"{year}-{int(mon)+1:02d}-01"
        # Adjust end_date to last day of month
        end_date = (datetime.strptime(end_date, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
    elif days:
        end_date = datetime.now().strftime('%Y-%m-%d')
        start_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    else:
        # Default: last 30 days
        end_date = datetime.now().strftime('%Y-%m-%d')
        start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')

    return start_date, end_date


# ===== TMDB DISCOVERY FUNCTIONS =====

def discover_movies_from_tmdb(month: Optional[str] = None, days: Optional[int] = None) -> list:
    """Query TMDB directly for movies with premiere dates in the specified period."""
    movies = []

    if month:
        year, mon = month.split('-')
        start_date = f"{month}-01"
        if mon == '12':
            end_date = f"{int(year)+1}-01-01"
        else:
            end_date = f"{year}-{int(mon)+1:02d}-01"
    elif days:
        end_date = datetime.now().strftime('%Y-%m-%d')
        start_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    else:
        end_date = datetime.now().strftime('%Y-%m-%d')
        start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')

    logger.info(f"Querying TMDB for premieres: {start_date} to {end_date}")

    page = 1
    total_pages = 1

    while page <= total_pages and page <= 500:
        try:
            resp = requests.get(f"{TMDB_BASE}/discover/movie", params={
                'api_key': TMDB_API_KEY,
                'primary_release_date.gte': start_date,
                'primary_release_date.lte': end_date,
                'sort_by': 'primary_release_date.desc',
                'page': page
            }, timeout=15)

            if resp.status_code != 200:
                logger.warning(f"TMDB returned {resp.status_code}")
                break

            data = resp.json()
            total_pages = data.get('total_pages', 1)

            for movie in data.get('results', []):
                movies.append((str(movie['id']), movie.get('release_date', '')))

            if page == 1:
                logger.info(f"Found {data.get('total_results', 0)} movies across {total_pages} pages")

            page += 1
            time.sleep(0.05)

        except Exception as e:
            logger.error(f"Error on page {page}: {e}")
            break

    return movies


def get_release_dates(tmdb_id: str) -> dict:
    """Get release dates from TMDB API."""
    if not TMDB_API_KEY:
        return {}
    try:
        url = f"{TMDB_BASE}/movie/{tmdb_id}/release_dates"
        resp = requests.get(url, params={'api_key': TMDB_API_KEY}, timeout=10)
        if resp.status_code == 200:
            return resp.json()
    except:
        pass
    return {}


def get_watch_providers(tmdb_id: str) -> dict:
    """Get watch providers from TMDB API."""
    if not TMDB_API_KEY:
        return {}
    try:
        url = f"{TMDB_BASE}/movie/{tmdb_id}/watch/providers"
        resp = requests.get(url, params={'api_key': TMDB_API_KEY}, timeout=10)
        if resp.status_code == 200:
            return resp.json()
    except:
        pass
    return {}


def get_movie_details(tmdb_id: str) -> dict:
    """Get movie details including credits and videos."""
    if not TMDB_API_KEY:
        return {}
    try:
        url = f"{TMDB_BASE}/movie/{tmdb_id}"
        resp = requests.get(url, params={
            'api_key': TMDB_API_KEY,
            'append_to_response': 'credits,videos'
        }, timeout=10)
        if resp.status_code == 200:
            return resp.json()
    except:
        pass
    return {}


def has_type4_release(release_data: dict) -> tuple:
    """Check if movie has Type 4 (Digital) release. Returns (has_type4, date)."""
    results = release_data.get('results', [])
    for country in results:
        for release in country.get('release_dates', []):
            if release.get('type') == 4:
                return True, release.get('release_date', '')[:10]
    return False, ''


def has_watch_providers(provider_data: dict) -> bool:
    """Check if movie has any watch providers (streaming, rent, buy)."""
    results = provider_data.get('results', {})
    for country_code in ['US'] + list(results.keys()):
        if country_code not in results:
            continue
        country_data = results[country_code]
        if country_data.get('flatrate') or country_data.get('rent') or country_data.get('buy'):
            return True
    return False


def get_director(details: dict) -> str:
    """Extract director from credits."""
    credits = details.get('credits', {})
    crew = credits.get('crew', [])
    for person in crew:
        if person.get('job') == 'Director':
            return person.get('name', '')
    return ''


def get_trailer_url(details: dict) -> str:
    """Extract YouTube trailer URL from videos."""
    videos = details.get('videos', {}).get('results', [])
    # Prefer official trailers, then teasers, then any video
    for video_type in ['Trailer', 'Teaser', 'Clip']:
        for video in videos:
            if video.get('type') == video_type and video.get('site') == 'YouTube':
                return f"https://www.youtube.com/watch?v={video.get('key')}"
    # Fallback to any YouTube video
    for video in videos:
        if video.get('site') == 'YouTube':
            return f"https://www.youtube.com/watch?v={video.get('key')}"
    return ''


def is_wrestling(movie: dict) -> bool:
    """Check if movie is a wrestling event (to filter out)."""
    title_lower = movie.get('title', '').lower()
    overview_lower = movie.get('overview', '').lower()
    return any(term in title_lower or term in overview_lower for term in
               ['wrestling', 'wwe', 'aew', 'njpw', 'stardom', 'ddt', 'roh'])


def classify_drc_platform(homepage_url: str) -> str:
    """Classify DRC platform type from homepage URL."""
    if not homepage_url:
        return ''

    try:
        parsed = urlparse(homepage_url)
        domain = parsed.netloc.lower()
        if domain.startswith('www.'):
            domain = domain[4:]

        # MAJOR PLATFORMS TO EXCLUDE
        exclusions = {
            'netflix.com': 'EXCLUDED_netflix',
            'tv.apple.com': 'EXCLUDED_apple',
            'apple.com/tv': 'EXCLUDED_apple',
            'itunes.apple.com': 'EXCLUDED_apple',
            'amazon.com': 'EXCLUDED_amazon',
            'primevideo.com': 'EXCLUDED_amazon',
            'hulu.com': 'EXCLUDED_hulu',
            'disneyplus.com': 'EXCLUDED_disney',
            'hbomax.com': 'EXCLUDED_hbo',
            'max.com': 'EXCLUDED_hbo',
            'paramountplus.com': 'EXCLUDED_paramount',
            'peacocktv.com': 'EXCLUDED_peacock',
            'vudu.com': 'EXCLUDED_vudu',
            'play.google.com': 'EXCLUDED_google',
        }

        for pattern, result in exclusions.items():
            if pattern in domain:
                return result

        # VALID DRC PLATFORMS
        platforms = {
            'vimeo.com': 'vimeo',
            'youtube.com': 'youtube',
            'youtu.be': 'youtube',
            'patreon.com': 'patreon',
            'letterboxd.com': 'letterboxd',
            'substack.com': 'substack',
            'gumroad.com': 'gumroad',
            'fourthwall.com': 'fourthwall',
            'filmhub.com': 'filmhub',
            'vhx.tv': 'vhx',
            'uscreen.io': 'uscreen',
            'eventive.org': 'eventive',
            'seedandspark.com': 'seed_and_spark',
        }

        for pattern, result in platforms.items():
            if pattern in domain:
                return result

        return 'filmmaker_site'
    except Exception:
        return 'unknown'


def categorize_movie(movie: dict) -> str:
    """Categorize movie by runtime only."""
    runtime = movie.get('runtime')
    if runtime is None:
        return 'Unknown Runtime'
    if runtime >= 70:
        return 'Feature Films (70+ min)'
    if runtime >= 30:
        return 'Medium (30-69 min)'
    if runtime >= 10:
        return 'Shorts (10-29 min)'
    return 'Micro (<10 min)'


def sort_by_platform_and_runtime(gems: list) -> list:
    """Sort gems by platform priority then runtime (descending)."""
    platform_priority = {
        'vimeo': 1, 'youtube': 2, 'patreon': 3, 'letterboxd': 4,
        'substack': 5, 'gumroad': 6, 'fourthwall': 7, 'filmhub': 8,
        'filmmaker_site': 9, 'unknown': 10, '': 11
    }

    def sort_key(gem):
        platform = gem.get('drc_platform', '')
        priority = platform_priority.get(platform, 99)
        runtime = gem.get('runtime') or 0
        return (priority, -runtime)

    return sorted(gems, key=sort_key)


def find_hidden_gems(movies: list, verbose: bool = True) -> list:
    """Find movies with Type 4 release but no watch providers."""
    hidden_gems = []
    total = len(movies)

    for i, (tmdb_id, intake_date) in enumerate(movies):
        if verbose and (i + 1) % 50 == 0:
            logger.info(f"Processed {i + 1}/{total}...")

        release_data = get_release_dates(tmdb_id)
        has_type4, type4_date = has_type4_release(release_data)

        if not has_type4:
            time.sleep(0.05)
            continue

        provider_data = get_watch_providers(tmdb_id)
        if has_watch_providers(provider_data):
            time.sleep(0.05)
            continue

        details = get_movie_details(tmdb_id)
        if not details:
            time.sleep(0.05)
            continue

        genre_names = [g.get('name', '') for g in details.get('genres', [])]

        gem = {
            'id': tmdb_id,
            'title': details.get('title', 'Unknown'),
            'intake_date': intake_date,
            'type4_date': type4_date,
            'runtime': details.get('runtime'),
            'overview': details.get('overview', ''),
            'homepage': details.get('homepage', ''),
            'genres': genre_names,
            'director': get_director(details),
            'trailer_url': get_trailer_url(details),
            'imdb': details.get('imdb_id')
        }

        gem['drc_platform'] = classify_drc_platform(gem['homepage'])

        if gem['drc_platform'].startswith('EXCLUDED_'):
            time.sleep(0.05)
            continue

        if is_wrestling(gem):
            time.sleep(0.05)
            continue

        if gem.get('runtime') is None or gem.get('runtime') < 70:
            time.sleep(0.05)
            continue

        gem['category'] = categorize_movie(gem)
        gem['add_to_nrw'] = False
        hidden_gems.append(gem)

        time.sleep(0.05)

    return hidden_gems


# ===== UNIFIED SCRAPING FUNCTIONS (using standalone scrapers) =====

def discover_from_patreon(max_creators: int = 20, min_runtime: int = 10, headless: bool = True,
                         start_date: str = None, end_date: str = None) -> List[Dict]:
    """Discover films from Patreon creators using standalone scraper."""
    if not PATREON_AVAILABLE:
        logger.warning("Patreon scraper not available, skipping")
        return []

    logger.info("Discovering films from Patreon...")

    try:
        scraper = PatreonFilmScraper(headless=headless)
        films = scraper.discover_films(max_creators=max_creators)
        scraper.cleanup()

        # Convert to gems format
        gems = []
        for film in films:
            # Extract post date if available
            post_date = film.get('post_date') or film.get('discovery_date')

            # Apply date filtering if specified
            if start_date and end_date and post_date:
                if not is_within_date_range(post_date, start_date, end_date):
                    continue

            gems.append({
                'id': film.get('id'),
                'title': film.get('title', ''),
                'intake_date': film.get('discovery_date', datetime.now().strftime('%Y-%m-%d')),
                'type4_date': '',
                'runtime': film.get('runtime'),
                'overview': film.get('description', ''),
                'homepage': film.get('post_url', ''),
                'genres': film.get('genres', []),
                'director': film.get('creator_name', ''),
                'imdb': None,
                'drc_platform': 'patreon',
                'category': categorize_movie(film),
                'patreon_url': film.get('post_url', ''),
                'creator_name': film.get('creator_name', ''),
                'creator_username': film.get('creator_username', ''),
                'post_date': post_date,
                'tmdb_match_confidence': film.get('tmdb_match_confidence', 'none'),
                'add_to_nrw': False
            })

        logger.info(f"Patreon discovery: {len(gems)} films")
        return gems

    except Exception as e:
        logger.error(f"Error in Patreon discovery: {e}")
        return []


def discover_from_vimeo(max_pages: int = 5, min_runtime: int = 70, headless: bool = True,
                       start_date: str = None, end_date: str = None) -> List[Dict]:
    """Discover films from Vimeo On Demand using standalone scraper."""
    if not VIMEO_AVAILABLE:
        logger.warning("Vimeo scraper not available, skipping")
        return []

    logger.info("Discovering films from Vimeo On Demand...")

    try:
        scraper = VimeoOnDemandScraper(headless=headless)
        films = scraper.discover_films(max_pages=max_pages)
        scraper.cleanup()

        # Convert to gems format
        gems = []
        for film in films:
            if film.get('runtime', 0) < min_runtime:
                continue

            # Extract upload date if available
            upload_date = film.get('upload_date') or film.get('discovery_date')

            # Apply date filtering if specified
            if start_date and end_date and upload_date:
                if not is_within_date_range(upload_date, start_date, end_date):
                    continue

            gems.append({
                'id': film.get('id'),
                'title': film.get('title', ''),
                'intake_date': film.get('discovery_date', datetime.now().strftime('%Y-%m-%d')),
                'type4_date': '',
                'runtime': film.get('runtime'),
                'overview': film.get('description', '') or film.get('tmdb_overview', ''),
                'homepage': film.get('vimeo_url', ''),
                'genres': film.get('genres', []),
                'director': film.get('filmmaker', ''),
                'imdb': None,
                'drc_platform': 'vimeo',
                'category': categorize_movie(film),
                'vimeo_url': film.get('vimeo_url', ''),
                'price_rent': film.get('price_rent'),
                'price_buy': film.get('price_buy'),
                'upload_date': upload_date,
                'tmdb_match_confidence': film.get('tmdb_match_confidence', 'none'),
                'add_to_nrw': False
            })

        logger.info(f"Vimeo discovery: {len(gems)} films")
        return gems

    except Exception as e:
        logger.error(f"Error in Vimeo discovery: {e}")
        return []


def discover_from_youtube(max_pages: int = 5, max_searches: int = None, min_runtime: int = 70, headless: bool = True,
                         start_date: str = None, end_date: str = None) -> List[Dict]:
    """Discover films from YouTube using standalone scraper."""
    if not YOUTUBE_AVAILABLE:
        logger.warning("YouTube scraper not available, skipping")
        return []

    logger.info("Discovering films from YouTube...")

    # Use max_searches if provided, otherwise max_pages
    pages = max_searches if max_searches is not None else max_pages

    try:
        scraper = YouTubePaidScraper(headless=headless)
        films = scraper.discover_films(max_pages=pages)
        scraper.cleanup()

        # Convert to gems format
        gems = []
        for film in films:
            if film.get('runtime', 0) < min_runtime:
                continue

            # Extract upload date if available
            upload_date = film.get('upload_date') or film.get('discovery_date')

            # Apply date filtering if specified
            if start_date and end_date and upload_date:
                if not is_within_date_range(upload_date, start_date, end_date):
                    continue

            gems.append({
                'id': film.get('id'),
                'title': film.get('title', ''),
                'intake_date': film.get('discovery_date', datetime.now().strftime('%Y-%m-%d')),
                'type4_date': '',
                'runtime': film.get('runtime'),
                'overview': film.get('description', '') or film.get('tmdb_overview', ''),
                'homepage': film.get('youtube_url', ''),
                'genres': film.get('genres', []),
                'director': film.get('channel_name', ''),
                'imdb': None,
                'drc_platform': 'youtube',
                'category': categorize_movie(film),
                'youtube_url': film.get('youtube_url', ''),
                'channel_name': film.get('channel_name', ''),
                'is_free': film.get('is_free', False),
                'price_rent': film.get('price_rent'),
                'price_buy': film.get('price_buy'),
                'upload_date': upload_date,
                'tmdb_match_confidence': film.get('tmdb_match_confidence', 'none'),
                'add_to_nrw': False
            })

        logger.info(f"YouTube discovery: {len(gems)} films")
        return gems

    except Exception as e:
        logger.error(f"Error in YouTube discovery: {e}")
        return []


def discover_from_letterboxd(sections: List[str] = None, headless: bool = True) -> List[Dict]:
    """Discover films from Letterboxd Video Store using standalone scraper."""
    if not LETTERBOXD_AVAILABLE:
        logger.warning("Letterboxd scraper not available, skipping")
        return []

    logger.info("Discovering films from Letterboxd Video Store...")

    try:
        scraper = LetterboxdScraper()
        films = scraper.discover_films(sections=sections)

        # Convert to gems format
        gems = []
        for film in films:
            gems.append({
                'id': film.get('tmdb_id'),
                'title': film.get('title', ''),
                'intake_date': film.get('discovery_date', datetime.now().strftime('%Y-%m-%d')),
                'type4_date': '',
                'runtime': film.get('runtime'),
                'overview': film.get('tmdb_overview', ''),
                'homepage': film.get('letterboxd_url', ''),
                'genres': [],
                'director': film.get('director', ''),
                'imdb': None,
                'drc_platform': 'letterboxd',
                'category': categorize_movie(film) if film.get('runtime') else 'Unknown Runtime',
                'letterboxd_url': film.get('letterboxd_url', ''),
                'letterboxd_slug': film.get('letterboxd_slug', ''),
                'rent_link': film.get('rent_link', ''),
                'tmdb_match_confidence': film.get('tmdb_match_confidence', 'none'),
                'add_to_nrw': False
            })

        logger.info(f"Letterboxd discovery: {len(gems)} films")
        return gems

    except Exception as e:
        logger.error(f"Error in Letterboxd discovery: {e}")
        return []


# ===== OUTPUT FUNCTIONS =====

def save_csv(gems: list, output_path: str):
    """Save hidden gems to CSV."""
    category_order = [
        'Feature Films (70+ min)', 'Medium (30-69 min)',
        'Shorts (10-29 min)', 'Micro (<10 min)', 'Unknown Runtime'
    ]

    def sort_key(g):
        cat_idx = category_order.index(g['category']) if g['category'] in category_order else 99
        return (cat_idx, g.get('title', '').lower())

    sorted_gems = sorted(gems, key=sort_key)

    with open(output_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Title', 'Category', 'DRC Platform', 'Runtime', 'Genres', 'Director',
                        'Watch Link', 'IMDB', 'TMDB', 'Overview', 'Price (Rent)', 'Price (Buy)', 'TMDB Confidence'])

        for gem in sorted_gems:
            imdb_url = f"https://www.imdb.com/title/{gem['imdb']}/" if gem.get('imdb') else ""
            tmdb_id = gem.get('id', '')
            tmdb_url = f"https://www.themoviedb.org/movie/{tmdb_id}" if tmdb_id else ""

            writer.writerow([
                gem.get('title', ''),
                gem.get('category', ''),
                gem.get('drc_platform', ''),
                gem.get('runtime', ''),
                ', '.join(gem.get('genres', [])),
                gem.get('director', ''),
                gem.get('homepage', ''),
                imdb_url,
                tmdb_url,
                gem.get('overview', '')[:150],
                gem.get('price_rent', ''),
                gem.get('price_buy', ''),
                gem.get('tmdb_match_confidence', '')
            ])

    logger.info(f"Saved {len(gems)} hidden gems to {output_path}")


def save_excel_with_tabs(results_by_source: Dict[str, List[Dict]], output_path: str,
                        start_date: str = None, end_date: str = None):
    """Save results to Excel with separate tabs per source and a metrics summary."""
    if not EXCEL_AVAILABLE:
        print("⚠️  openpyxl not available, skipping Excel export")
        return

    wb = Workbook()

    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14)

    # Create Metrics sheet first
    ws_metrics = wb.active
    ws_metrics.title = "Metrics"

    # Add title and date range at top
    ws_metrics.append(["Hidden Gems Scraper Results"])
    ws_metrics.cell(row=1, column=1).font = title_font
    if start_date and end_date:
        ws_metrics.append([f"Date Range: {start_date} to {end_date}"])
    else:
        ws_metrics.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws_metrics.append([])  # Blank row

    # Metrics header (now starting at row 4)
    ws_metrics.append(["Source", "Films Found", "With Runtime", "With Date", "Avg Runtime"])
    for col in range(1, 6):
        cell = ws_metrics.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill

    # Calculate metrics for each source
    total_films = 0
    for source, gems in results_by_source.items():
        count = len(gems)
        total_films += count
        with_runtime = sum(1 for g in gems if g.get('runtime'))
        with_date = sum(1 for g in gems if g.get('upload_date') or g.get('post_date'))
        runtimes = [g.get('runtime', 0) for g in gems if g.get('runtime')]
        avg_runtime = round(sum(runtimes) / len(runtimes), 1) if runtimes else 0

        ws_metrics.append([source, count, with_runtime, with_date, avg_runtime])

    # Total row
    ws_metrics.append([])
    ws_metrics.append(["TOTAL", total_films, "", "", ""])
    ws_metrics.cell(row=ws_metrics.max_row, column=1).font = Font(bold=True)

    # Adjust column widths for metrics
    ws_metrics.column_dimensions['A'].width = 20
    ws_metrics.column_dimensions['B'].width = 15
    ws_metrics.column_dimensions['C'].width = 15
    ws_metrics.column_dimensions['D'].width = 15
    ws_metrics.column_dimensions['E'].width = 15

    # Create a tab for each source
    for source, gems in results_by_source.items():
        if not gems:
            continue

        # Clean sheet name (Excel has restrictions)
        sheet_name = source[:31].replace('/', '-').replace('\\', '-')
        ws = wb.create_sheet(title=sheet_name)

        # Determine columns based on source
        if source == "TMDB":
            headers = ["Title", "Runtime", "Director", "Genres", "Homepage", "Trailer", "TMDB ID", "IMDB", "Overview"]
        elif source == "Vimeo":
            headers = ["Title", "Runtime", "Upload Date", "Vimeo URL", "Description", "Price"]
        elif source == "YouTube":
            headers = ["Title", "Runtime", "Upload Date", "Channel", "YouTube URL", "Duration Text"]
        elif source == "Patreon":
            headers = ["Title", "Runtime", "Post Date", "Creator", "Patreon URL", "Description"]
        else:
            headers = ["Title", "Runtime", "URL", "Description"]

        # Write headers
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for gem in gems:
            if source == "TMDB":
                row = [
                    gem.get('title', ''),
                    gem.get('runtime', ''),
                    gem.get('director', ''),
                    ', '.join(gem.get('genres', [])) if gem.get('genres') else '',
                    gem.get('homepage', ''),
                    gem.get('trailer_url', ''),
                    gem.get('id', ''),
                    gem.get('imdb', ''),
                    (gem.get('overview', '') or '')[:200]
                ]
            elif source == "Vimeo":
                row = [
                    gem.get('title', ''),
                    gem.get('runtime', ''),
                    gem.get('upload_date', ''),
                    gem.get('vimeo_url', ''),
                    (gem.get('description', '') or '')[:200],
                    gem.get('price_text', '')
                ]
            elif source == "YouTube":
                row = [
                    gem.get('title', ''),
                    gem.get('runtime', ''),
                    gem.get('upload_date', ''),
                    gem.get('channel_name', ''),
                    gem.get('youtube_url', ''),
                    gem.get('duration_text', '')
                ]
            elif source == "Patreon":
                row = [
                    gem.get('title', ''),
                    gem.get('runtime', ''),
                    gem.get('post_date', ''),
                    gem.get('creator_name', ''),
                    gem.get('patreon_url', ''),
                    (gem.get('description', '') or '')[:200]
                ]
            else:
                row = [
                    gem.get('title', ''),
                    gem.get('runtime', ''),
                    gem.get('homepage', '') or gem.get('url', ''),
                    (gem.get('description', '') or gem.get('overview', '') or '')[:200]
                ]
            ws.append(row)

            # Make URLs clickable hyperlinks
            row_num = ws.max_row
            if source == "TMDB":
                # Homepage link (column E)
                homepage = gem.get('homepage', '')
                if homepage:
                    ws.cell(row=row_num, column=5).hyperlink = homepage
                    ws.cell(row=row_num, column=5).style = "Hyperlink"
                # Trailer link (column F)
                trailer = gem.get('trailer_url', '')
                if trailer:
                    ws.cell(row=row_num, column=6).hyperlink = trailer
                    ws.cell(row=row_num, column=6).style = "Hyperlink"
                # IMDB link (column H)
                imdb = gem.get('imdb', '')
                if imdb:
                    imdb_url = f"https://www.imdb.com/title/{imdb}/"
                    ws.cell(row=row_num, column=8).hyperlink = imdb_url
                    ws.cell(row=row_num, column=8).style = "Hyperlink"
            elif source in ["Vimeo", "YouTube", "Patreon"]:
                # URL column (column D for Vimeo/YouTube, column E for Patreon)
                url_col = 5 if source == "Patreon" else 4
                url = row[url_col - 1] if len(row) >= url_col else ''
                if url:
                    ws.cell(row=row_num, column=url_col).hyperlink = url
                    ws.cell(row=row_num, column=url_col).style = "Hyperlink"

        # Adjust column widths based on source
        if source == "TMDB":
            ws.column_dimensions['A'].width = 40  # Title
            ws.column_dimensions['B'].width = 8   # Runtime
            ws.column_dimensions['C'].width = 25  # Director
            ws.column_dimensions['D'].width = 20  # Genres
            ws.column_dimensions['E'].width = 45  # Homepage
            ws.column_dimensions['F'].width = 45  # Trailer
            ws.column_dimensions['G'].width = 12  # TMDB ID
            ws.column_dimensions['H'].width = 14  # IMDB
            ws.column_dimensions['I'].width = 60  # Overview
        else:
            ws.column_dimensions['A'].width = 40  # Title
            ws.column_dimensions['B'].width = 8   # Runtime
            for col_letter in ['C', 'D', 'E', 'F', 'G', 'H']:
                if col_letter in ws.column_dimensions:
                    ws.column_dimensions[col_letter].width = 25

    wb.save(output_path)
    print(f"📊 Excel saved: {output_path}")
    print(f"   Tabs: {', '.join(results_by_source.keys())}")


def save_curation_csv(gems: list, output_path: str):
    """Save hidden gems in curation-ready CSV format."""
    sorted_gems = sort_by_platform_and_runtime(gems)

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Add to NRW', 'Title', 'Platform', 'Runtime', 'Director/Channel',
            'Watch Link', 'Price (Rent)', 'Price (Buy)', 'TMDB ID', 'IMDB',
            'Genres', 'Overview', 'Discovery Source', 'TMDB Confidence'
        ])

        for gem in sorted_gems:
            # Determine discovery source
            discovery_source = 'TMDB'
            if gem.get('vimeo_url'):
                discovery_source = 'Vimeo Scraper'
            elif gem.get('youtube_url'):
                discovery_source = 'YouTube Scraper'
            elif gem.get('patreon_url'):
                discovery_source = 'Patreon Scraper'
            elif gem.get('letterboxd_url'):
                discovery_source = 'Letterboxd Scraper'

            director = gem.get('director', '') or gem.get('channel_name', '') or gem.get('filmmaker', '')

            platform_names = {
                'vimeo': 'Vimeo', 'youtube': 'YouTube', 'patreon': 'Patreon',
                'letterboxd': 'Letterboxd', 'substack': 'Substack', 'gumroad': 'Gumroad',
                'fourthwall': 'Fourthwall', 'filmhub': 'FilmHub',
                'filmmaker_site': 'Filmmaker Site', 'unknown': 'Unknown', '': 'No Homepage'
            }
            platform = platform_names.get(gem.get('drc_platform', ''), gem.get('drc_platform', ''))

            watch_link = (gem.get('vimeo_url') or gem.get('youtube_url') or
                         gem.get('patreon_url') or gem.get('letterboxd_url') or gem.get('homepage', ''))

            tmdb_id = gem.get('id', '')
            tmdb_url = f"https://www.themoviedb.org/movie/{tmdb_id}" if tmdb_id else ""
            imdb_id = gem.get('imdb', '')
            imdb_url = f"https://www.imdb.com/title/{imdb_id}/" if imdb_id else ""

            price_rent = gem.get('price_rent', '')
            price_buy = gem.get('price_buy', '')
            if gem.get('is_free'):
                price_rent = 'Free'
                price_buy = 'Free'

            writer.writerow([
                'FALSE',
                gem.get('title', ''),
                platform,
                gem.get('runtime', ''),
                director,
                watch_link,
                price_rent,
                price_buy,
                tmdb_url,
                imdb_url,
                ', '.join(gem.get('genres', [])),
                gem.get('overview', '')[:200],
                discovery_source,
                gem.get('tmdb_match_confidence', '')
            ])

    logger.info(f"Saved curation CSV to {output_path}")


def save_curation_json(gems: list, output_path: str):
    """Save hidden gems in curation-ready JSON format."""
    sorted_gems = sort_by_platform_and_runtime(gems)

    curation_data = {
        'generated_at': datetime.now().isoformat(),
        'total_films': len(sorted_gems),
        'films': []
    }

    for gem in sorted_gems:
        discovery_source = 'TMDB'
        if gem.get('vimeo_url'):
            discovery_source = 'Vimeo Scraper'
        elif gem.get('youtube_url'):
            discovery_source = 'YouTube Scraper'
        elif gem.get('patreon_url'):
            discovery_source = 'Patreon Scraper'
        elif gem.get('letterboxd_url'):
            discovery_source = 'Letterboxd Scraper'

        director = gem.get('director', '') or gem.get('channel_name', '') or gem.get('filmmaker', '')
        watch_link = (gem.get('vimeo_url') or gem.get('youtube_url') or
                     gem.get('patreon_url') or gem.get('letterboxd_url') or gem.get('homepage', ''))

        curation_film = {
            'add_to_nrw': False,
            'title': gem.get('title', ''),
            'tmdb_id': gem.get('id'),
            'platform': gem.get('drc_platform', ''),
            'watch_url': watch_link,
            'runtime': gem.get('runtime'),
            'director': director,
            'description': gem.get('overview', ''),
            'price': {
                'rent': gem.get('price_rent'),
                'buy': gem.get('price_buy'),
                'is_free': gem.get('is_free', False)
            },
            'discovery_source': discovery_source,
            'tmdb_confidence': gem.get('tmdb_match_confidence', 'none'),
            'genres': gem.get('genres', []),
            'release_year': gem.get('release_year'),
            'imdb_id': gem.get('imdb'),
            'metadata': {
                'intake_date': gem.get('intake_date', ''),
                'type4_date': gem.get('type4_date', ''),
                'discovery_date': gem.get('discovery_date', ''),
                'category': gem.get('category', '')
            }
        }
        curation_data['films'].append(curation_film)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(curation_data, f, indent=2, ensure_ascii=False)

    logger.info(f"Saved curation JSON to {output_path}")


def save_markdown(gems: list, output_path: str):
    """Save hidden gems to markdown."""
    category_order = [
        'Feature Films (70+ min)', 'Medium (30-69 min)',
        'Shorts (10-29 min)', 'Micro (<10 min)', 'Unknown Runtime'
    ]

    by_category = {}
    for gem in gems:
        cat = gem.get('category', 'Unknown')
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append(gem)

    md = f"# Hidden Gems - {datetime.now().strftime('%Y-%m-%d')}\n"
    md += f"{len(gems)} indie films with digital releases not on major platforms\n\n"
    md += "---\n\n"

    for cat_name in category_order:
        if cat_name not in by_category:
            continue

        cat_gems = sorted(by_category[cat_name], key=lambda x: x.get('title', '').lower())
        md += f"## {cat_name} ({len(cat_gems)})\n\n"

        for gem in cat_gems:
            title = gem.get('title', 'Unknown')
            runtime = gem.get('runtime')
            runtime_str = f"{runtime}min" if runtime else ""
            genres = gem.get('genres', [])[:2]
            genre_str = ', '.join(genres) if genres else ''
            director = gem.get('director', '')
            homepage = gem.get('homepage', '')
            drc_platform = gem.get('drc_platform', '')

            platform_labels = {
                'vimeo': 'Vimeo', 'youtube': 'YouTube', 'patreon': 'Patreon',
                'letterboxd': 'Letterboxd', 'substack': 'Substack', 'gumroad': 'Gumroad',
                'filmmaker_site': 'Filmmaker Site'
            }
            platform_label = platform_labels.get(drc_platform, '')

            info_parts = [p for p in [platform_label, runtime_str, genre_str, director] if p]
            info = " · ".join(info_parts) if info_parts else ""

            links = []
            if homepage:
                links.append(f"[Watch]({homepage})")
            if gem.get('imdb'):
                links.append(f"[IMDB](https://www.imdb.com/title/{gem['imdb']}/)")
            link_str = " · ".join(links)

            if info:
                md += f"- **{title}** — {info}\n  {link_str}\n"
            else:
                md += f"- **{title}**\n  {link_str}\n"

        md += "\n"

    with open(output_path, 'w') as f:
        f.write(md)

    logger.info(f"Saved markdown to {output_path}")


def save_json(gems: list, output_path: str):
    """Save raw JSON for future use."""
    with open(output_path, 'w') as f:
        json.dump(gems, f, indent=2)
    logger.info(f"Saved JSON to {output_path}")


def print_summary(gems: list, curation_mode: bool = False):
    """Print category summary."""
    by_category = {}
    for gem in gems:
        cat = gem.get('category', 'Unknown')
        by_category[cat] = by_category.get(cat, 0) + 1

    print(f"\n📊 Found {len(gems)} hidden gems:")
    for cat, count in sorted(by_category.items(), key=lambda x: -x[1]):
        print(f"   {cat}: {count}")

    with_watch = len([g for g in gems if g.get('homepage')])
    with_imdb = len([g for g in gems if g.get('imdb')])
    print(f"\n   {with_watch} have direct watch links")
    print(f"   {with_imdb} have IMDB pages")

    by_platform = {}
    for gem in gems:
        platform = gem.get('drc_platform', 'unknown') or 'no_homepage'
        by_platform[platform] = by_platform.get(platform, 0) + 1

    print(f"\nPlatform breakdown:")
    platform_names = {
        'vimeo': 'Vimeo', 'youtube': 'YouTube', 'patreon': 'Patreon',
        'letterboxd': 'Letterboxd', 'substack': 'Substack', 'gumroad': 'Gumroad',
        'filmmaker_site': 'Filmmaker sites', 'unknown': 'Unknown', 'no_homepage': 'No homepage'
    }
    for platform, count in sorted(by_platform.items(), key=lambda x: -x[1]):
        name = platform_names.get(platform, platform)
        print(f"   {name}: {count}")

    if curation_mode:
        print("\n📋 Curation Workflow:")
        print("   1. Open the *_curation.csv file")
        print("   2. Review each film's watch link and metadata")
        print("   3. Change 'FALSE' to 'TRUE' in 'Add to NRW' column for films to add")


def main():
    parser = argparse.ArgumentParser(description='Find hidden gem indie films')
    parser.add_argument('--month', help='Filter by month (e.g., 2025-12)')
    parser.add_argument('--days', type=int, default=30, help='Last N days (default: 30)')
    parser.add_argument('--output', default='hidden_gems', help='Output filename prefix')
    parser.add_argument('--quiet', action='store_true', help='Suppress progress output')

    # Live scraping options
    parser.add_argument('--scrape-vimeo', action='store_true', help='Scrape Vimeo On Demand')
    parser.add_argument('--scrape-youtube', action='store_true', help='Scrape YouTube Movies')
    parser.add_argument('--scrape-patreon', action='store_true', help='Scrape Patreon creators')
    parser.add_argument('--scrape-letterboxd', action='store_true', help='Scrape Letterboxd Video Store')
    parser.add_argument('--scrape-all', action='store_true', help='Scrape all platforms')

    # Scraping configuration
    parser.add_argument('--max-creators', type=int, default=10, help='Max Patreon creators (default: 10)')
    parser.add_argument('--max-pages', type=int, default=3, help='Max pages per platform (default: 3)')
    parser.add_argument('--min-runtime', type=int, default=70, help='Minimum runtime (default: 70)')
    parser.add_argument('--headless', action='store_true', default=True, help='Headless mode (default: true)')
    parser.add_argument('--no-headless', action='store_false', dest='headless', help='Run with GUI')

    parser.add_argument('--export-for-curation', action='store_true',
                       help='Export in curation-ready format')
    args = parser.parse_args()

    # Configure logging
    log_level = logging.WARNING if args.quiet else logging.INFO
    logging.basicConfig(
        level=log_level,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

    if not TMDB_API_KEY:
        print("❌ TMDB_API_KEY environment variable not set")
        return 1

    # Generate dated, numbered output path
    output_prefix = get_dated_output_prefix()
    print(f"📁 Output: {output_prefix}.*")

    print("🎬 Hidden Gems Scraper")
    print("=" * 40)

    # Query TMDB for movies with premiere dates
    if args.month:
        print(f"🔍 Discovering movies from {args.month}...")
        movies = discover_movies_from_tmdb(month=args.month)
    else:
        print(f"🔍 Discovering movies from last {args.days} days...")
        movies = discover_movies_from_tmdb(days=args.days)

    print(f"   Found {len(movies)} movies to check")

    # Calculate date range (used for filtering and Excel header)
    start_date, end_date = calculate_date_range(month=args.month, days=args.days)

    # Track results by source for Excel export
    results_by_source = {}

    gems = []
    if movies:
        print(f"\n🔎 Checking for hidden gems (Type 4 release, no providers)...")
        tmdb_gems = find_hidden_gems(movies, verbose=not args.quiet)
        results_by_source['TMDB'] = tmdb_gems
        gems.extend(tmdb_gems)

    # DIRECT SCRAPING using standalone scrapers
    scrape_vimeo = args.scrape_vimeo or args.scrape_all
    scrape_youtube = args.scrape_youtube or args.scrape_all
    scrape_patreon = args.scrape_patreon or args.scrape_all
    scrape_letterboxd = args.scrape_letterboxd or args.scrape_all

    if scrape_vimeo or scrape_youtube or scrape_patreon or scrape_letterboxd:
        print(f"\n🚀 LIVE SCRAPING MODE")
        print("=" * 40)
        print(f"   Date range: {start_date} to {end_date}")

        if scrape_vimeo:
            vimeo_gems = discover_from_vimeo(
                max_pages=args.max_pages,
                min_runtime=args.min_runtime,
                headless=args.headless,
                start_date=start_date,
                end_date=end_date
            )
            results_by_source['Vimeo'] = vimeo_gems
            gems.extend(vimeo_gems)

        if scrape_youtube:
            youtube_gems = discover_from_youtube(
                max_searches=args.max_pages,
                min_runtime=args.min_runtime,
                headless=args.headless,
                start_date=start_date,
                end_date=end_date
            )
            results_by_source['YouTube'] = youtube_gems
            gems.extend(youtube_gems)

        if scrape_patreon:
            patreon_gems = discover_from_patreon(
                max_creators=args.max_creators,
                min_runtime=args.min_runtime,
                headless=args.headless,
                start_date=start_date,
                end_date=end_date
            )
            results_by_source['Patreon'] = patreon_gems
            gems.extend(patreon_gems)

        if scrape_letterboxd:
            letterboxd_gems = discover_from_letterboxd(headless=args.headless)
            results_by_source['Letterboxd'] = letterboxd_gems
            gems.extend(letterboxd_gems)

        # Deduplicate by title
        seen_titles = set()
        unique_gems = []
        for gem in gems:
            title_key = gem.get('title', '').lower().strip()
            if title_key and title_key not in seen_titles:
                seen_titles.add(title_key)
                unique_gems.append(gem)
        gems = unique_gems
        print(f"\n   After deduplication: {len(gems)} total films")

    if not gems:
        print("😢 No hidden gems found")
        return 0

    # Save outputs
    print()
    if args.export_for_curation:
        save_curation_csv(gems, f"{output_prefix}_curation.csv")
        save_curation_json(gems, f"{output_prefix}_curation.json")
        print("\n📋 Curation exports generated!")
    else:
        save_csv(gems, f"{output_prefix}.csv")
        save_markdown(gems, f"{output_prefix}.md")
        save_json(gems, f"{output_prefix}.json")

    # Always save Excel with tabs (for test visibility)
    if results_by_source:
        save_excel_with_tabs(results_by_source, f"{output_prefix}.xlsx", start_date, end_date)

    print_summary(gems, curation_mode=args.export_for_curation)

    return 0


if __name__ == '__main__':
    exit(main())
